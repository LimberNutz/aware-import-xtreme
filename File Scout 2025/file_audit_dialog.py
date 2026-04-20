"""
File Audit Dialog - Google Drive Inspector for Traveler Files
Integrated into File Scout application
"""
import sys
import os
import pickle
import webbrowser
import concurrent.futures
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QGridLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QCheckBox,
    QProgressBar, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QRadioButton, QButtonGroup, QTextEdit, QComboBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor

# Google Drive & Sheets API imports
try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False

# API Scopes
SCOPES = [
    'https://www.googleapis.com/auth/drive.readonly',
    'https://www.googleapis.com/auth/spreadsheets.readonly'
]

import openpyxl
import re
from urllib.parse import urlparse, parse_qs


class GoogleSheetsParser:
    """Parser for Google Sheets traveler files"""
    
    def __init__(self, sheet_id, sheets_service, tabs):
        self.sheet_id = sheet_id
        self.service = sheets_service
        self.tabs = tabs
    
    def parse_all(self):
        """Parse all specified tabs from Google Sheet"""
        all_entities = []
        
        for sheet_name in self.tabs:
            try:
                entities = self.parse_sheet(sheet_name)
                all_entities.extend(entities)
            except Exception as e:
                print(f"Warning: Could not parse sheet '{sheet_name}': {e}")
                continue
        
        return all_entities
    
    def get_column_mapping(self, sheet_name):
        """Get column indices for each field based on sheet type
        
        Different traveler sheets have different column layouts.
        Returns dict mapping field names to column indices (0-based).
        """
        # API-510 Traveler columns (starts at H)
        if 'API-510' in sheet_name or '510' in sheet_name:
            return {
                'API': 7,           # Column H
                'EXT VT DATE': 8,   # Column I
                'EXT VT REPORT': 9, # Column J
                'INT VT DATE': 10,  # Column K
                'INT VT REPORT': 11,# Column L
                'TECH': 12,         # Column M
                'UT DATE': 13,      # Column N
                'UT REPORT': 14,    # Column O
                'U1A': 15,          # Column P
                'DWG BY': 16,       # Column Q
                'DWG DATE': 17,     # Column R
                # DR fields not in 510
                'DR': 999,
                'DR BY': 999,
                'DR DATE': 999,
                'DR REPORT': 999,
            }
        
        # API-570 Traveler columns (starts at G!)
        elif 'API-570' in sheet_name or '570' in sheet_name:
            return {
                'API': 6,           # Column G
                'EXT VT DATE': 7,   # Column H (VT DATE)
                'EXT VT REPORT': 8, # Column I (VT REPORT)
                'INT VT DATE': 999, # Not present in 570
                'INT VT REPORT': 999, # Not present in 570
                'TECH': 9,          # Column J
                'UT DATE': 10,      # Column K
                'UT REPORT': 11,    # Column L
                'DWG BY': 12,       # Column M
                'DWG DATE': 13,     # Column N
                'DR': 14,           # Column O (Y/N if DR required)
                'DR BY': 15,        # Column P (initials)
                'DR DATE': 16,      # Column Q (date)
                'DR REPORT': 17,    # Column R (report completion)
                'U1A': 999,         # Not in 570
            }
        
        # Tank Traveler columns (starts at G!)
        elif 'Tank' in sheet_name:
            return {
                'API': 6,           # Column G
                'EXT VT DATE': 7,   # Column H (VT DATE)
                'EXT VT REPORT': 8, # Column I
                'INT VT DATE': 9,   # Column J (INTERNAL DATE)
                'INT VT REPORT': 10,# Column K
                'TECH': 11,         # Column L
                'UT DATE': 12,      # Column M
                'UT REPORT': 13,    # Column N
                'DWG BY': 14,       # Column O
                'DWG DATE': 15,     # Column P
                # DR fields not in Tank
                'DR': 999,
                'DR BY': 999,
                'DR DATE': 999,
                'DR REPORT': 999,
                'U1A': 999,
            }
        
        # Default to API-510 layout if unknown
        else:
            return {
                'API': 7, 'EXT VT DATE': 8, 'EXT VT REPORT': 9,
                'INT VT DATE': 10, 'INT VT REPORT': 11, 'TECH': 12,
                'UT DATE': 13, 'UT REPORT': 14,
            }
    
    def parse_sheet(self, sheet_name):
        """Parse a single sheet from Google Sheets"""
        entities = []
        
        # Read data from row 7 onwards (data starts at row 7, headers at row 6)
        range_name = f"'{sheet_name}'!A7:O1000"  # Read up to row 1000
        
        try:
            result = self.service.spreadsheets().values().get(
                spreadsheetId=self.sheet_id,
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            
            if not values:
                return entities
            
            # Also get hyperlinks using the full sheet data
            sheet_data = self.service.spreadsheets().get(
                spreadsheetId=self.sheet_id,
                ranges=[f"'{sheet_name}'!A7:A1000"],
                fields='sheets(data(rowData(values(hyperlink,formattedValue))))'
            ).execute()
            
            # Extract hyperlinks from column A
            hyperlinks = {}
            try:
                rows = sheet_data['sheets'][0]['data'][0].get('rowData', [])
                for idx, row in enumerate(rows):
                    if 'values' in row and len(row['values']) > 0:
                        cell = row['values'][0]
                        if 'hyperlink' in cell:
                            hyperlinks[idx] = cell['hyperlink']
            except (KeyError, IndexError):
                pass
            
            # Parse each row
            for row_idx, row in enumerate(values):
                if not row or len(row) == 0:
                    continue
                
                entity_name = row[0] if len(row) > 0 else ''
                if not entity_name or str(entity_name).strip() == '':
                    continue
                
                entity_name = str(entity_name).strip()
                actual_row = 7 + row_idx  # Convert to 1-indexed row number
                
                # Get folder URL from hyperlinks
                folder_url = hyperlinks.get(row_idx)
                folder_id = self.extract_folder_id(folder_url) if folder_url else None
                
                # Get sheet-specific column mapping
                cols = self.get_column_mapping(sheet_name)
                
                # Build entity data with correct column indices for this sheet type
                entity_data = {
                    'tab': sheet_name,
                    'row': actual_row,
                    'entity': entity_name,
                    'folder_url': folder_url,
                    'folder_id': folder_id,
                    'API': row[cols['API']] if len(row) > cols['API'] else None,
                    'EXT VT DATE': row[cols['EXT VT DATE']] if len(row) > cols['EXT VT DATE'] else None,
                    'EXT VT REPORT': row[cols['EXT VT REPORT']] if len(row) > cols['EXT VT REPORT'] else None,
                    'INT VT DATE': row[cols['INT VT DATE']] if len(row) > cols['INT VT DATE'] and cols['INT VT DATE'] < 999 else None,
                    'INT VT REPORT': row[cols['INT VT REPORT']] if len(row) > cols['INT VT REPORT'] and cols['INT VT REPORT'] < 999 else None,
                    'TECH': row[cols['TECH']] if len(row) > cols['TECH'] else None,
                    'UT DATE': row[cols['UT DATE']] if len(row) > cols['UT DATE'] else None,
                    'UT REPORT': row[cols['UT REPORT']] if len(row) > cols['UT REPORT'] else None,
                    'DWG BY': row[cols['DWG BY']] if len(row) > cols['DWG BY'] and cols['DWG BY'] < 999 else None,
                    'DWG DATE': row[cols['DWG DATE']] if len(row) > cols['DWG DATE'] and cols['DWG DATE'] < 999 else None,
                    'DR': row[cols['DR']] if len(row) > cols['DR'] and cols['DR'] < 999 else None,
                    'DR BY': row[cols['DR BY']] if len(row) > cols['DR BY'] and cols['DR BY'] < 999 else None,
                    'DR DATE': row[cols['DR DATE']] if len(row) > cols['DR DATE'] and cols['DR DATE'] < 999 else None,
                    'DR REPORT': row[cols['DR REPORT']] if len(row) > cols['DR REPORT'] and cols['DR REPORT'] < 999 else None,
                }
                
                entities.append(entity_data)
        
        except Exception as e:
            error_msg = str(e)
            if '403' in error_msg:
                raise Exception(
                    f"Google Sheets API access denied (403 error).\n\n"
                    f"Please enable the Google Sheets API:\n"
                    f"1. Visit: https://console.cloud.google.com/apis/library/sheets.googleapis.com\n"
                    f"2. Click 'Enable'\n"
                    f"3. Wait a few minutes and try again.\n\n"
                    f"Original error: {error_msg}"
                )
            print(f"Error parsing sheet {sheet_name}: {e}")
            raise
        
        return entities
    
    def extract_folder_id(self, url):
        """Extract folder ID from Google Drive URL"""
        if not url:
            return None
        
        patterns = [
            r'folders/([a-zA-Z0-9_-]+)',
            r'id=([a-zA-Z0-9_-]+)',
            r'/d/([a-zA-Z0-9_-]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None


class FileAuditWorker(QThread):
    """Background worker thread for running file audits"""
    
    progress = pyqtSignal(str)
    entity_progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, drive_service, sheets_service, traveler_input, tabs, initials_filter, parent_folder_url=None):
        super().__init__()
        self.drive_service = drive_service
        self.sheets_service = sheets_service
        self.traveler_input = traveler_input
        self.tabs = tabs
        self.initials_filter = initials_filter
        self.parent_folder_url = parent_folder_url
        self.missing_items = []
        self.folder_lookup = {}  # Phase 1: entity_name -> folder_id mapping
        
    def run(self):
        """Run the audit in background"""
        try:
            # Parse traveler
            self.progress.emit("📂 Parsing traveler file...")
            entities = self.parse_traveler()
            
            if not entities:
                self.error.emit("No entities found in traveler file")
                return
            
            self.progress.emit(f"✅ Found {len(entities)} entities")
            
            # PHASE 1: Build folder lookup from parent folder (10-20x speedup)
            if self.parent_folder_url:
                self.progress.emit("⚡ Building folder lookup from parent folder...")
                self.build_folder_lookup(self.parent_folder_url, entities)
                self.progress.emit(f"✅ Found {len(self.folder_lookup)} folders in parent")
            
            # Audit each entity
            total = len(entities)
            
            # PHASE 2: Batch file listing for speed (ALWAYS do this if any folders found)
            # Get all folder IDs that need checking
            folder_ids = [e['folder_id'] for e in entities if e.get('folder_id')]
            
            if folder_ids:
                self.progress.emit(f"⚡ Batch loading {len(folder_ids)} folder contents...")
                # Batch list files for all folders at once
                folder_files_cache = self.batch_list_files(folder_ids)
                self.progress.emit(f"✅ Loaded contents for {len(folder_files_cache)} folders")
            else:
                folder_files_cache = None
                self.progress.emit("⚠️ No folder IDs found - check if traveler has Drive links")
            
            # Sequential entity auditing (with cached folder contents)
            cache_hits = 0
            cache_misses = 0
            
            for idx, entity_data in enumerate(entities, 1):
                entity_name = entity_data['entity']
                
                # Pass cached folder files if available
                if folder_files_cache and entity_data.get('folder_id') in folder_files_cache:
                    entity_data['_cached_files'] = folder_files_cache[entity_data['folder_id']]
                    cache_hits += 1
                    self.entity_progress.emit(idx, total, f"⚡ {entity_name} (cached)")
                else:
                    cache_misses += 1
                    self.entity_progress.emit(idx, total, f"🔍 {entity_name} (scanning...)")
                
                try:
                    missing = self.audit_entity(entity_data)
                    self.missing_items.extend(missing)
                except Exception as e:
                    # Log error but continue with other entities
                    import traceback
                    error_detail = traceback.format_exc()
                    print(f"Error auditing {entity_name}: {error_detail}")
                    self.entity_progress.emit(idx, total, f"❌ {entity_name} (Error: {str(e)})")
            
            # Report cache effectiveness
            if folder_files_cache:
                self.progress.emit(f"📊 Cache stats: {cache_hits} hits, {cache_misses} misses")
            
            self.progress.emit(f"✅ Audit complete. Found {len(self.missing_items)} missing items.")
            self.finished.emit(self.missing_items)
            
        except Exception as e:
            import traceback
            error_msg = f"Audit error: {str(e)}\n\n{traceback.format_exc()}"
            self.error.emit(error_msg)
    
    def detect_input_type(self):
        """Detect if input is a Google Sheet URL/ID or local Excel file"""
        input_text = self.traveler_input.strip()
        
        # Check for Google Sheets URL
        if 'docs.google.com/spreadsheets' in input_text:
            return 'google_sheet', self.extract_sheet_id(input_text)
        
        # Check for direct Sheet ID (33+ alphanumeric characters)
        if len(input_text) > 30 and re.match(r'^[a-zA-Z0-9_-]+$', input_text):
            return 'google_sheet', input_text
        
        # Check for local file
        if Path(input_text).exists() and input_text.endswith(('.xlsx', '.xls')):
            return 'excel_file', input_text
        
        return 'unknown', None
    
    def extract_sheet_id(self, url):
        """Extract Sheet ID from Google Sheets URL"""
        # Pattern: https://docs.google.com/spreadsheets/d/{SHEET_ID}/...
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
        if match:
            return match.group(1)
        return None
    
    def parse_traveler(self):
        """Parse traveler from either Google Sheets or Excel file"""
        input_type, identifier = self.detect_input_type()
        
        if input_type == 'google_sheet':
            self.progress.emit("📊 Reading from live Google Sheet...")
            return self.parse_google_sheet(identifier)
        elif input_type == 'excel_file':
            self.progress.emit("📂 Reading from Excel file...")
            return self.parse_excel_file(identifier)
        else:
            raise ValueError("Invalid input: Must be a Google Sheet URL/ID or local Excel file path")
    
    def parse_google_sheet(self, sheet_id):
        """Parse traveler from Google Sheets"""
        parser = GoogleSheetsParser(sheet_id, self.sheets_service, self.tabs)
        entities = parser.parse_all()
        
        # Apply initials filter
        if self.initials_filter:
            entities = self.filter_by_initials(entities)
        
        return entities
    
    def parse_excel_file(self, file_path):
        """Parse traveler from local Excel file"""
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        all_entities = []
        
        for sheet_name in self.tabs:
            if sheet_name not in workbook.sheetnames:
                continue
            
            sheet = workbook[sheet_name]
            
            # Parse entities from sheet (headers in row 6, data starts row 7)
            for row_num in range(7, sheet.max_row + 1):
                entity_cell = sheet[f'A{row_num}']
                entity_name = entity_cell.value
                
                if not entity_name or str(entity_name).strip() == '':
                    continue
                
                entity_name = str(entity_name).strip()
                
                # Get Drive folder link
                folder_url = None
                folder_id = None
                if entity_cell.hyperlink:
                    folder_url = entity_cell.hyperlink.target
                    folder_id = self.extract_folder_id(folder_url)
                
                # Get sheet-specific column mapping
                cols = self.get_column_mapping(sheet_name)
                
                # Helper to convert column index to Excel letter
                def col_letter(idx):
                    """Convert 0-based column index to Excel letter (7 -> 'H')"""
                    if idx >= 999:  # Column doesn't exist for this sheet type
                        return None
                    return chr(65 + idx)  # 65 is 'A'
                
                # Extract column values using correct columns for this sheet type
                entity_data = {
                    'tab': sheet_name,
                    'row': row_num,
                    'entity': entity_name,
                    'folder_url': folder_url,
                    'folder_id': folder_id,
                    'API': sheet[f'{col_letter(cols["API"])}{row_num}'].value if col_letter(cols['API']) else None,
                    'EXT VT DATE': sheet[f'{col_letter(cols["EXT VT DATE"])}{row_num}'].value if col_letter(cols['EXT VT DATE']) else None,
                    'EXT VT REPORT': sheet[f'{col_letter(cols["EXT VT REPORT"])}{row_num}'].value if col_letter(cols['EXT VT REPORT']) else None,
                    'INT VT DATE': sheet[f'{col_letter(cols["INT VT DATE"])}{row_num}'].value if col_letter(cols['INT VT DATE']) else None,
                    'INT VT REPORT': sheet[f'{col_letter(cols["INT VT REPORT"])}{row_num}'].value if col_letter(cols['INT VT REPORT']) else None,
                    'TECH': sheet[f'{col_letter(cols["TECH"])}{row_num}'].value if col_letter(cols['TECH']) else None,
                    'UT DATE': sheet[f'{col_letter(cols["UT DATE"])}{row_num}'].value if col_letter(cols['UT DATE']) else None,
                    'UT REPORT': sheet[f'{col_letter(cols["UT REPORT"])}{row_num}'].value if col_letter(cols['UT REPORT']) else None,
                    'DWG BY': sheet[f'{col_letter(cols["DWG BY"])}{row_num}'].value if col_letter(cols['DWG BY']) else None,
                    'DWG DATE': sheet[f'{col_letter(cols["DWG DATE"])}{row_num}'].value if col_letter(cols['DWG DATE']) else None,
                    'DR': sheet[f'{col_letter(cols["DR"])}{row_num}'].value if col_letter(cols['DR']) else None,
                    'DR BY': sheet[f'{col_letter(cols["DR BY"])}{row_num}'].value if col_letter(cols['DR BY']) else None,
                    'DR DATE': sheet[f'{col_letter(cols["DR DATE"])}{row_num}'].value if col_letter(cols['DR DATE']) else None,
                    'DR REPORT': sheet[f'{col_letter(cols["DR REPORT"])}{row_num}'].value if col_letter(cols['DR REPORT']) else None,
                }
                
                # Filter by initials if specified
                if self.initials_filter:
                    filter_upper = self.initials_filter.upper()
                    api_val = str(entity_data.get('API', '')).upper()
                    tech_val = str(entity_data.get('TECH', '')).upper()
                    
                    if filter_upper not in api_val and filter_upper not in tech_val:
                        continue
                
                all_entities.append(entity_data)
        
        return all_entities
    
    def filter_by_initials(self, entities):
        """Filter entities by inspector initials"""
        if not self.initials_filter:
            return entities
        
        filter_upper = self.initials_filter.upper()
        filtered = []
        
        for entity_data in entities:
            api_val = str(entity_data.get('API', '')).upper()
            tech_val = str(entity_data.get('TECH', '')).upper()
            
            if filter_upper in api_val or filter_upper in tech_val:
                filtered.append(entity_data)
        
        return filtered
    
    def extract_folder_id(self, url):
        """Extract folder ID from Google Drive URL"""
        if not url:
            return None
        
        patterns = [
            r'folders/([a-zA-Z0-9_-]+)',
            r'id=([a-zA-Z0-9_-]+)',
            r'/d/([a-zA-Z0-9_-]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None
    
    def build_folder_lookup(self, parent_folder_url, entities):
        """PHASE 1: Build folder lookup from parent folder for 10-20x speedup
        
        Lists all subfolders from parent once, then maps entity names to folder IDs.
        This replaces individual hyperlink following with a single batch operation.
        """
        try:
            # Extract parent folder ID
            parent_id = self.extract_folder_id(parent_folder_url)
            if not parent_id:
                self.progress.emit("⚠️ Could not extract parent folder ID, falling back to hyperlinks")
                return
            
            # List all folders in parent (PHASE 2: Can be batched further if needed)
            query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.drive_service.files().list(
                q=query,
                fields='files(id, name)',
                pageSize=1000,  # Get up to 1000 folders at once
                supportsAllDrives=True,
                includeItemsFromAllDrives=True
            ).execute()
            
            folders = results.get('files', [])
            
            # Handle pagination if more than 1000 folders
            while 'nextPageToken' in results:
                results = self.drive_service.files().list(
                    q=query,
                    fields='files(id, name)',
                    pageSize=1000,
                    pageToken=results['nextPageToken'],
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True
                ).execute()
                folders.extend(results.get('files', []))
            
            # Build lookup dictionary: folder_name -> folder_id
            folder_dict = {folder['name']: folder['id'] for folder in folders}
            
            # Map entity names to folder IDs
            for entity in entities:
                entity_name = entity['entity']
                if entity_name in folder_dict:
                    # Found via parent lookup - override hyperlink
                    entity['folder_id'] = folder_dict[entity_name]
                    entity['folder_url'] = f"https://drive.google.com/drive/folders/{folder_dict[entity_name]}"
                    self.folder_lookup[entity_name] = folder_dict[entity_name]
            
            self.progress.emit(f"⚡ Mapped {len(self.folder_lookup)}/{len(entities)} entities to folders")
            
        except Exception as e:
            self.progress.emit(f"⚠️ Parent folder lookup failed: {str(e)}, using hyperlinks")
    
    def audit_entity(self, entity_data):
        """Audit a single entity for missing files"""
        missing = []
        
        # Get assignee info early for better tracking
        api_init = entity_data.get('API', '')
        tech_init = entity_data.get('TECH', '')
        primary_assignee = api_init if self.is_value_present(api_init) else (tech_init if self.is_value_present(tech_init) else 'N/A')
        
        # Skip if no folder link
        if not entity_data.get('folder_id'):
            missing.append({
                'Timestamp': datetime.now().strftime('%m/%d/%Y %H:%M:%S'),
                'Tab': entity_data['tab'],
                'Row': entity_data['row'],
                'Entity': entity_data['entity'],
                'FolderId': '',
                'FolderUrl': entity_data.get('folder_url', ''),
                'Task': 'Folder Link',
                'Assignee': primary_assignee,  # Now assigns to actual inspector
                'Trigger': f'API: {api_init}, TECH: {tech_init}' if self.is_value_present(api_init) or self.is_value_present(tech_init) else 'No inspector assigned',
                'Result': 'MISSING',
                'Notes': 'Missing entity hyperlink to Drive folder'
            })
            return missing
        
        folder_id = entity_data['folder_id']
        
        # PHASE 2: Use cached folder files if available (from batch loading)
        if '_cached_files' in entity_data:
            files = entity_data['_cached_files']
            # DEBUG: Confirm cache hit
            # print(f"✅ Cache HIT for {entity_data['entity']} ({len(files)} files)")
        else:
            # Fall back to individual folder scan (slower)
            # print(f"⚠️ Cache MISS for {entity_data['entity']} - scanning individually")
            files = self.list_files_recursive(folder_id)
        
        # Differentiate between access denied and empty folder
        if not files:
            # Try to determine if it's an access issue or truly empty
            try:
                # Attempt a direct folder query
                test_query = self.drive_service.files().get(
                    fileId=folder_id,
                    fields='id,name',
                    supportsAllDrives=True
                ).execute()
                # If we got here, we can access the folder - it's just empty
                notes = 'Folder is empty (no files or subfolders found)'
            except Exception:
                # Access denied or folder doesn't exist
                notes = 'Unable to access folder (permission denied or folder deleted)'
            
            missing.append({
                'Timestamp': datetime.now().strftime('%m/%d/%Y %H:%M:%S'),
                'Tab': entity_data['tab'],
                'Row': entity_data['row'],
                'Entity': entity_data['entity'],
                'FolderId': folder_id,
                'FolderUrl': entity_data.get('folder_url', ''),
                'Task': 'Folder Access',
                'Assignee': primary_assignee,
                'Trigger': '',
                'Result': 'MISSING',
                'Notes': notes
            })
            return missing
        ext_vt_date = entity_data.get('EXT VT DATE', '')
        ext_vt_report = entity_data.get('EXT VT REPORT', '')
        ut_report = entity_data.get('UT REPORT', '')
        
        # Check for photos if VT date is filled
        if self.is_value_present(api_init) and self.is_value_present(ext_vt_date):
            photo_result = self.check_photos(files)
            if not photo_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'Photos', api_init, f'Date: {ext_vt_date}',
                    photo_result['reason']
                ))
        
        # Check for External VT report if report column has completion date
        # EXT VT REPORT column gets filled with a DATE when report is uploaded
        if self.is_value_present(api_init) and self.is_value_present(ext_vt_report):
            vt_ext_result = self.check_vt_external(files)
            if not vt_ext_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'EXT VT Report', api_init, f'Report Date: {ext_vt_report}',
                    vt_ext_result['reason']
                ))
        
        # Check for Internal VT report if report column has completion date
        # INT VT REPORT column gets filled with a DATE when report is uploaded
        int_vt_date = entity_data.get('INT VT DATE', '')
        int_vt_report = entity_data.get('INT VT REPORT', '')
        if self.is_value_present(api_init) and self.is_value_present(int_vt_report):
            vt_int_result = self.check_vt_internal(files)
            if not vt_int_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'INT VT Report', api_init, f'Report Date: {int_vt_report}',
                    vt_int_result['reason']
                ))
        
        # Check for UT report if report column has completion date
        # UT REPORT column gets filled with a DATE when report is uploaded
        if self.is_value_present(tech_init) and self.is_value_present(ut_report):
            ut_result = self.check_ut_report(files)
            if not ut_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'UT Report', tech_init, f'Report Date: {ut_report}',
                    ut_result['reason']
                ))
        
        # Check for DR (Digital Radiography) - API-570 only
        # Check when DR DATE is filled (work completed)
        dr_by = entity_data.get('DR BY', '')
        dr_date = entity_data.get('DR DATE', '')
        
        if self.is_value_present(dr_date):
            # DR work is completed (DR DATE is filled), check for DR files
            assignee = dr_by if self.is_value_present(dr_by) else 'N/A'
            dr_result = self.check_dr_report(files)
            if not dr_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'DR Report', assignee, f'DR Date: {dr_date}',
                    dr_result['reason']
                ))
        
        # Check for Field Sketch (DWG column) - All tabs
        # Check when DWG DATE is filled (sketch completed)
        dwg_by = entity_data.get('DWG BY', '')
        dwg_date = entity_data.get('DWG DATE', '')
        
        if self.is_value_present(dwg_date):
            # Field sketch is completed (DWG DATE is filled), check for sketch files
            assignee = dwg_by if self.is_value_present(dwg_by) else 'N/A'
            dwg_result = self.check_dwg_files(files)
            if not dwg_result['found']:
                missing.append(self.create_missing_item(
                    entity_data, 'Field Sketch', assignee, f'Sketch Date: {dwg_date}',
                    dwg_result['reason']
                ))
        
        return missing
    
    def list_files_recursive(self, folder_id):
        """List all files in a folder recursively"""
        return self._list_files_recursive_with_service(folder_id, self.drive_service)
    
    def _list_files_recursive_with_service(self, folder_id, drive_service):
        """List all files in a folder recursively using a specific drive service instance
        
        This allows for thread-safe parallel scanning by using separate service instances.
        """
        all_files = []
        
        def recurse(fid, path=''):
            try:
                query = f"'{fid}' in parents and trashed=false"
                results = drive_service.files().list(
                    q=query,
                    fields='files(id, name, mimeType)',
                    pageSize=1000,
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True
                ).execute()
                
                items = results.get('files', [])
                
                for item in items:
                    is_folder = item['mimeType'] == 'application/vnd.google-apps.folder'
                    current_path = f"{path}/{item['name']}" if path else item['name']
                    
                    all_files.append({
                        'name': item['name'],
                        'id': item['id'],
                        'path': current_path.lower(),
                        'is_folder': is_folder,
                        'mime_type': item['mimeType']
                    })
                    
                    if is_folder:
                        recurse(item['id'], current_path)
            
            except Exception as e:
                pass  # Skip inaccessible folders
        
        recurse(folder_id)
        return all_files
    
    def batch_list_files(self, folder_ids):
        """PHASE 2: Parallel folder scanning (10x speedup without Batch API)
        
        Uses multi-threading to scan multiple folders concurrently.
        This works even if Google Batch API is not available/enabled.
        Returns dict: {folder_id: [files]}
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        
        results = {}
        results_lock = threading.Lock()
        completed = 0
        total = len(folder_ids)
        
        # IMPORTANT: Google API client is NOT thread-safe
        # Each thread needs its own Drive service instance
        def scan_folder(folder_id):
            """Scan a single folder recursively with thread-local Drive service"""
            try:
                # Create a thread-local Drive service instance
                # This is necessary because the Google API client is NOT thread-safe
                from googleapiclient.discovery import build
                thread_drive_service = build('drive', 'v3', credentials=self.drive_service._http.credentials)
                
                # Use thread-local service to list files
                files = self._list_files_recursive_with_service(folder_id, thread_drive_service)
                return folder_id, files, None
            except Exception as e:
                import traceback
                error_msg = f"{str(e)}\n{traceback.format_exc()}"
                return folder_id, [], error_msg
        
        # Use thread pool to scan folders in parallel (10 concurrent threads)
        max_workers = 10
        self.progress.emit(f"⚡ Scanning {total} folders with {max_workers} parallel threads...")
        
        try:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all folder scans
                future_to_folder = {executor.submit(scan_folder, fid): fid for fid in folder_ids}
                
                # Process results as they complete
                for future in as_completed(future_to_folder):
                    try:
                        folder_id, files, error = future.result()
                        
                        with results_lock:
                            results[folder_id] = files
                            completed += 1
                            
                            # Progress update every 10 folders
                            if completed % 10 == 0 or completed == total:
                                self.progress.emit(f"  ⚡ Scanned {completed}/{total} folders ({len(files)} files in last folder)")
                    except Exception as e:
                        self.progress.emit(f"⚠️ Error processing folder result: {str(e)}")
                        continue
            
            self.progress.emit(f"✅ Parallel scan complete: {len(results)} folders scanned")
        except Exception as e:
            self.progress.emit(f"⚠️ Thread pool error: {str(e)}")
            import traceback
            print(f"Thread pool failed:")
            print(traceback.format_exc())
            # Fall back to sequential scanning
            self.progress.emit("Falling back to sequential scanning...")
            for folder_id in folder_ids:
                results[folder_id] = self.list_files_recursive(folder_id)
        
        return results
    
    def is_value_present(self, value):
        """Check if a value is present (not None, not N/A, not empty) - handles various data types"""
        if value is None:
            return False
        
        # Handle datetime objects from Excel
        if hasattr(value, 'strftime'):
            return True
        
        # Handle numeric values
        if isinstance(value, (int, float)):
            return True
        
        # Handle string values
        str_value = str(value).strip().upper()
        return str_value != '' and str_value not in ['N/A', 'NA', 'NONE']
    
    def create_missing_item(self, entity_data, task, assignee, trigger, notes):
        """Create a missing item record"""
        return {
            'Timestamp': datetime.now().strftime('%m/%d/%Y %H:%M:%S'),
            'Tab': entity_data['tab'],
            'Row': entity_data['row'],
            'Entity': entity_data['entity'],
            'FolderId': entity_data.get('folder_id', ''),
            'FolderUrl': entity_data.get('folder_url', ''),  # Store the URL for hyperlinks
            'Task': task,
            'Assignee': assignee or '',
            'Trigger': trigger or '',
            'Result': 'MISSING',
            'Notes': notes or ''
        }
    
    def check_photos(self, files):
        """Check for inspection photos - either in dedicated folders or loose in entity folder
        
        Relaxed requirement: 3+ valid images = good to go (no keywords/folder required)
        """
        # Common image extensions
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif', '.webp', '.heic']
        
        # Files to exclude (non-inspection images)
        exclude_keywords = ['logo', 'icon', 'signature', 'letterhead', 'banner', 'header']
        
        # Diagnostic counters
        total_images = 0
        valid_images = 0  # Images that aren't logos/icons/etc
        
        for f in files:
            if f['is_folder']:
                continue
            
            name_lower = f['name'].lower()
            
            # Check if it's an image file
            is_image = any(name_lower.endswith(ext) for ext in image_extensions)
            if not is_image:
                continue
            
            total_images += 1
            
            # Skip excluded images (logos, icons, etc.)
            if any(keyword in name_lower for keyword in exclude_keywords):
                continue
            
            valid_images += 1
        
        # Relaxed requirement: If 3+ valid images exist, consider it good
        if valid_images >= 3:
            return {'found': True, 'reason': f'Found {valid_images} valid images (3+ threshold met)'}
        
        # Build diagnostic reason for failure
        if total_images == 0:
            reason = 'No image files found in folder'
        elif valid_images == 0:
            excluded = total_images
            reason = f'Found {total_images} images, but all were excluded (logo/icon/signature files)'
        else:
            reason = f'Found only {valid_images} valid images (need at least 3 for inspection photos)'
        
        return {'found': False, 'reason': reason}
    
    def check_vt_external(self, files):
        """Check for external VT report file in external inspection folders"""
        # Look for files in external folders that are likely VT reports
        external_folder_names = ['external']
        vt_keywords = ['vt', 'visual', 'inspection']
        report_extensions = ['.xlsx', '.xls', '.pdf', '.docx', '.doc']
        
        # Diagnostic counters
        total_reports = 0
        external_files = 0
        vt_named_files = 0
        
        for f in files:
            if f['is_folder']:
                continue
            
            path_lower = f['path'].lower()
            name_lower = f['name'].lower()
            
            # Check if it's a report file type
            is_report_type = any(name_lower.endswith(ext) for ext in report_extensions)
            if is_report_type:
                total_reports += 1
            
            # Check if file is in an external folder
            in_external_folder = any(folder in path_lower for folder in external_folder_names)
            if in_external_folder and is_report_type:
                external_files += 1
            
            # Check if filename suggests it's a VT report
            is_vt_file = any(keyword in name_lower for keyword in vt_keywords)
            if is_vt_file and is_report_type:
                vt_named_files += 1
            
            # Must be in external folder and be a VT-related report
            if in_external_folder and is_vt_file and is_report_type:
                return {'found': True, 'reason': f'Found external VT report: {f["name"]}'}
        
        # Build diagnostic reason
        if total_reports == 0:
            reason = 'No report files found in folder'
        elif external_files == 0:
            reason = f'Found {total_reports} reports, but none in EXTERNAL INSPECTION folders'
        elif vt_named_files == 0:
            reason = f'Found {external_files} reports in external folders, but none with VT/visual/inspection keywords'
        else:
            reason = f'Found {total_reports} reports, {external_files} in external folders, {vt_named_files} VT-related, but none matching all criteria'
        
        return {'found': False, 'reason': reason}
    
    def check_vt_internal(self, files):
        """Check for internal VT report file in internal inspection folders"""
        # Look for files in internal folders that are likely VT reports
        internal_folder_names = ['internal']
        vt_keywords = ['vt', 'visual', 'inspection', 'internal']
        report_extensions = ['.xlsx', '.xls', '.pdf', '.docx', '.doc']
        
        # Diagnostic counters
        total_reports = 0
        internal_files = 0
        vt_named_files = 0
        
        for f in files:
            if f['is_folder']:
                continue
            
            path_lower = f['path'].lower()
            name_lower = f['name'].lower()
            
            # Check if it's a report file type
            is_report_type = any(name_lower.endswith(ext) for ext in report_extensions)
            if is_report_type:
                total_reports += 1
            
            # Check if file is in an internal folder
            in_internal_folder = any(folder in path_lower for folder in internal_folder_names)
            if in_internal_folder and is_report_type:
                internal_files += 1
            
            # Check if filename suggests it's a VT report
            is_vt_file = any(keyword in name_lower for keyword in vt_keywords)
            if is_vt_file and is_report_type:
                vt_named_files += 1
            
            # Must be in internal folder and be a VT-related report
            if in_internal_folder and is_vt_file and is_report_type:
                return {'found': True, 'reason': f'Found internal VT report: {f["name"]}'}
        
        # Build diagnostic reason
        if total_reports == 0:
            reason = 'No report files found in folder'
        elif internal_files == 0:
            reason = f'Found {total_reports} reports, but none in INTERNAL INSPECTION folders'
        elif vt_named_files == 0:
            reason = f'Found {internal_files} reports in internal folders, but none with VT/visual/inspection keywords'
        else:
            reason = f'Found {total_reports} reports, {internal_files} in internal folders, {vt_named_files} VT-related, but none matching all criteria'
        
        return {'found': False, 'reason': reason}
    
    def check_ut_report(self, files):
        """Check for UT report file (thickness/UTT)"""
        # UT-related keywords
        ut_keywords = ['ut', 'thickness', 'utt', 'ultrasonic']
        # Report file extensions
        report_extensions = ['.xlsx', '.xls', '.pdf', '.docx', '.doc']
        
        # Diagnostic counters
        total_reports = 0
        ut_named_files = 0
        
        for f in files:
            if f['is_folder']:
                continue
            
            name_lower = f['name'].lower()
            
            # Check if it's a report file type
            is_report_type = any(name_lower.endswith(ext) for ext in report_extensions)
            if is_report_type:
                total_reports += 1
            
            # Check if filename contains UT keywords
            has_ut_keyword = any(keyword in name_lower for keyword in ut_keywords)
            if has_ut_keyword and is_report_type:
                return {'found': True, 'reason': f'Found UT report: {f["name"]}'}
            
            if has_ut_keyword:
                ut_named_files += 1
        
        # Build diagnostic reason
        if total_reports == 0:
            reason = 'No report files found in folder'
        elif ut_named_files == 0:
            reason = f'Found {total_reports} reports, but none with UT/thickness/ultrasonic keywords'
        else:
            reason = f'Found {ut_named_files} files with UT keywords, but wrong file types (need .xlsx, .xls, .pdf, .docx, .doc)'
        
        return {'found': False, 'reason': reason}
    
    def check_dr_report(self, files):
        """Check for DR (Digital Radiography) files - reports AND photos - API-570 only"""
        # DR-related keywords (Digital Radiography / Radiography Testing)
        dr_keywords = ['dr', 'rt', 'radiography', 'digital radiography']
        # Report file extensions
        report_extensions = ['.xlsx', '.xls', '.pdf', '.docx', '.doc']
        # Photo/Image extensions
        photo_extensions = ['.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp', '.heic']
        
        # Diagnostic counters
        total_reports = 0
        total_photos = 0
        dr_named_files = 0
        dr_photos_found = 0
        dr_reports_found = 0
        
        for f in files:
            if f['is_folder']:
                # Check if it's a DR-specific folder (e.g., "DR Photos", "DR", "Radiography")
                folder_lower = f['name'].lower()
                has_dr_keyword = any(keyword in folder_lower for keyword in dr_keywords)
                if has_dr_keyword:
                    # Found a DR folder - look for photos inside it
                    return {'found': True, 'reason': f'Found DR folder: {f["name"]}'}
                continue
            
            name_lower = f['name'].lower()
            
            # Check if it's a report file type
            is_report_type = any(name_lower.endswith(ext) for ext in report_extensions)
            if is_report_type:
                total_reports += 1
            
            # Check if it's a photo file type
            is_photo_type = any(name_lower.endswith(ext) for ext in photo_extensions)
            if is_photo_type:
                total_photos += 1
            
            # Check if filename contains DR keywords
            has_dr_keyword = any(keyword in name_lower for keyword in dr_keywords)
            
            if has_dr_keyword and is_report_type:
                dr_reports_found += 1
                return {'found': True, 'reason': f'Found DR report: {f["name"]}'}
            
            if has_dr_keyword and is_photo_type:
                dr_photos_found += 1
                return {'found': True, 'reason': f'Found DR photo: {f["name"]}'}
            
            if has_dr_keyword:
                dr_named_files += 1
        
        # Build diagnostic reason
        if total_reports == 0 and total_photos == 0:
            reason = 'No report or photo files found in folder'
        elif dr_named_files == 0:
            reason = f'Found {total_reports} reports and {total_photos} photos, but none with DR/RT/radiography keywords'
        else:
            reason = f'Found {dr_named_files} files with DR keywords, but wrong file types (need reports or images)'
        
        return {'found': False, 'reason': reason}
    
    def check_dwg_files(self, files):
        """Check for API Inspector's Field Sketch (DWG column in traveler)
        
        Looking for field sketches created by API inspector, NOT AutoCAD files (data team).
        Field sketches typically have ISO/isometric/sketch in the name.
        """
        # Field sketch keywords (what API inspectors name their sketches)
        sketch_keywords = ['iso', 'isometric', 'sketch', 'field sketch', 'field']
        # Field sketch file types (photos/PDFs of hand-drawn or digital sketches)
        sketch_extensions = ['.pdf', '.png', '.jpg', '.jpeg', '.tif', '.tiff', '.heic']
        
        # Diagnostic counters
        total_sketches = 0
        sketch_named_files = 0
        
        for f in files:
            if f['is_folder']:
                # Check if it's a sketch-specific folder
                folder_lower = f['name'].lower()
                has_sketch_keyword = any(keyword in folder_lower for keyword in sketch_keywords)
                if has_sketch_keyword:
                    # Found a sketch folder
                    return {'found': True, 'reason': f'Found sketch folder: {f["name"]}'}
                continue
            
            name_lower = f['name'].lower()
            
            # Check if it's a sketch file type (NOT .dwg/.dxf - those are data team files)
            is_sketch_type = any(name_lower.endswith(ext) for ext in sketch_extensions)
            if is_sketch_type:
                total_sketches += 1
            
            # Check if filename contains sketch keywords
            has_sketch_keyword = any(keyword in name_lower for keyword in sketch_keywords)
            
            if has_sketch_keyword and is_sketch_type:
                return {'found': True, 'reason': f'Found field sketch: {f["name"]}'}
            
            if has_sketch_keyword:
                sketch_named_files += 1
        
        # Build diagnostic reason
        if total_sketches == 0:
            reason = 'No sketch files found in folder'
        elif sketch_named_files == 0:
            reason = f'Found {total_sketches} image/PDF files, but none with ISO/sketch keywords'
        else:
            reason = f'Found {sketch_named_files} files with sketch keywords, but wrong file types (need .pdf or image files)'
        
        return {'found': False, 'reason': reason}


class FileAuditDialog(QDialog):
    """File Audit Dialog - Google Drive Inspector"""
    
    def __init__(self, parent=None, theme='light', zoom_level=100):
        super().__init__(parent)
        self.theme = theme
        self.zoom_level = zoom_level
        self.drive_service = None
        self.sheets_service = None
        self.worker = None
        self.missing_items = []
        
        self.setWindowTitle("File Audit - Google Drive & Sheets Inspector")
        self.setMinimumSize(900, 700)
        
        self._init_ui()
        self._apply_theme()
        
        # Apply zoom from parent window AFTER UI is created
        if zoom_level != 100:
            self._apply_zoom()
        
        # Check if Google API is available
        if not GOOGLE_API_AVAILABLE:
            self.show_warning(
                "Missing Dependencies",
                "Google Drive API libraries are not installed.\n\n"
                "Please install them with:\n"
                "pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client"
            )
    
    def _init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # File selection group
        file_group = QGroupBox("Traveler File")
        file_layout = QHBoxLayout()
        self.traveler_input = QLineEdit()
        self.traveler_input.setPlaceholderText("Excel file path OR Google Sheet URL...")
        self.traveler_input.setToolTip(
            "Enter:\n"
            "• Local Excel file path (.xlsx, .xls)\n"
            "• Google Sheet URL (https://docs.google.com/spreadsheets/d/...)\n"
            "• Google Sheet ID (direct ID string)"
        )
        browse_btn = QPushButton("Browse Local File...")
        browse_btn.clicked.connect(self.browse_traveler)
        file_layout.addWidget(self.traveler_input)
        file_layout.addWidget(browse_btn)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        # Parent folder group (PHASE 1: Speed optimization)
        parent_group = QGroupBox("Project Parent Folder (⚡ Speed Optimization)")
        parent_layout = QVBoxLayout()
        
        # Parent folder URL input
        parent_url_layout = QHBoxLayout()
        self.parent_folder_input = QLineEdit()
        self.parent_folder_input.setPlaceholderText("Google Drive parent folder URL (e.g., '2024 Inspection Field Report')...")
        self.parent_folder_input.setToolTip(
            "Enter the parent Google Drive folder URL that contains all entity subfolders.\n\n"
            "⚡ This enables 10-20x faster audits by listing all folders once\n"
            "instead of checking each entity hyperlink individually.\n\n"
            "Example: https://drive.google.com/drive/folders/ABC123...\n"
            "The parent folder containing: 3001-VE-001, 3002-VE-002, etc."
        )
        parent_url_layout.addWidget(QLabel("Parent Folder URL:"))
        parent_url_layout.addWidget(self.parent_folder_input)
        parent_layout.addLayout(parent_url_layout)
        
        # Info label
        info_label = QLabel(
            "ℹ️ Optional but recommended: Providing the parent folder URL speeds up audits by 10-20x.\n"
            "Leave blank to use individual entity hyperlinks (slower)."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666; font-size: 9pt; padding: 5px;")
        parent_layout.addWidget(info_label)
        
        parent_group.setLayout(parent_layout)
        layout.addWidget(parent_group)
        
        # Tab selection group
        tab_group = QGroupBox("Traveler Tabs to Audit")
        tab_layout = QHBoxLayout()
        self.tab_510 = QCheckBox("API-510 Traveler")
        self.tab_570 = QCheckBox("API-570 Traveler")
        self.tab_tank = QCheckBox("Tank Traveler")
        self.tab_510.setChecked(True)
        self.tab_570.setChecked(True)
        self.tab_tank.setChecked(True)
        tab_layout.addWidget(self.tab_510)
        tab_layout.addWidget(self.tab_570)
        tab_layout.addWidget(self.tab_tank)
        tab_layout.addStretch()
        tab_group.setLayout(tab_layout)
        layout.addWidget(tab_group)
        
        # Filter group
        filter_group = QGroupBox("Filters")
        filter_layout = QGridLayout()
        filter_layout.addWidget(QLabel("Inspector Initials:"), 0, 0)
        self.initials_input = QLineEdit()
        self.initials_input.setPlaceholderText("e.g., PO, CS (leave blank for all)")
        filter_layout.addWidget(self.initials_input, 0, 1)
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Control buttons
        control_layout = QHBoxLayout()
        self.auth_btn = QPushButton("🔐 Authenticate with Google")
        self.auth_btn.clicked.connect(self.authenticate_drive)
        self.start_btn = QPushButton("▶ Start Audit")
        self.start_btn.clicked.connect(self.start_audit)
        self.start_btn.setEnabled(False)
        control_layout.addWidget(self.auth_btn)
        control_layout.addWidget(self.start_btn)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("Ready. Please authenticate with Google Drive to begin.")
        layout.addWidget(self.status_label)
        
        # Results table
        results_group = QGroupBox("Audit Results")
        results_layout = QVBoxLayout()
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(10)
        self.results_table.setHorizontalHeaderLabels([
            "Timestamp", "Tab", "Row", "Entity", "Folder ID",
            "Task", "Assignee", "Trigger", "Result", "Notes"
        ])
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.results_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.results_table.cellClicked.connect(self.on_entity_clicked)  # Handle clicks on entity links
        self.results_table.cellEntered.connect(self.on_cell_hovered)  # Change cursor on hover
        self.results_table.setMouseTracking(True)  # Enable mouse tracking for hover effects
        results_layout.addWidget(self.results_table)
        
        results_group.setLayout(results_layout)
        layout.addWidget(results_group)
        
        # Export buttons
        export_layout = QHBoxLayout()
        self.export_excel_btn = QPushButton("📊 Export to Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setEnabled(False)
        self.export_csv_btn = QPushButton("📄 Export to CSV")
        self.export_csv_btn.clicked.connect(self.export_to_csv)
        self.export_csv_btn.setEnabled(False)
        export_layout.addWidget(self.export_excel_btn)
        export_layout.addWidget(self.export_csv_btn)
        export_layout.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        export_layout.addWidget(close_btn)
        layout.addLayout(export_layout)
    
    def _apply_theme(self):
        """Apply File Scout theme to dialog"""
        # Theme will be applied by parent when theme changes
        pass
    
    def _apply_zoom(self):
        """Apply zoom level to all dialog elements"""
        from PyQt6.QtWidgets import QWidget, QTableWidget, QGroupBox
        
        # Calculate new font size based on zoom
        base_font_size = 9
        new_font_size = int(base_font_size * (self.zoom_level / 100))
        
        # Create font for the dialog
        app_font = self.font()
        app_font.setPointSize(new_font_size)
        
        # Apply to dialog and all widgets
        self.setFont(app_font)
        
        # Update all child widgets
        for widget in self.findChildren(QWidget):
            widget.setFont(app_font)
        
        # Special handling for tables to adjust row heights
        for table in self.findChildren(QTableWidget):
            table.verticalHeader().setDefaultSectionSize(int(25 * (self.zoom_level / 100)))
        
        # Scale layout spacing
        zoom_factor = self.zoom_level / 100
        
        # Update all QGroupBox content margins
        for groupbox in self.findChildren(QGroupBox):
            layout = groupbox.layout()
            if layout:
                # Scale margins
                base_margin = 10
                scaled_margin = int(base_margin * zoom_factor)
                layout.setContentsMargins(scaled_margin, scaled_margin, scaled_margin, scaled_margin)
    
    def _create_message_box(self, icon, title, text, buttons=QMessageBox.StandardButton.Ok):
        """Create a QMessageBox with proper zoom-scaled font"""
        from PyQt6.QtWidgets import QPushButton
        msg_box = QMessageBox(icon, title, text, buttons, self)
        
        # Apply zoom to message box
        if self.zoom_level != 100:
            zoom_factor = self.zoom_level / 100
            font = msg_box.font()
            font.setPointSize(int(9 * zoom_factor))
            msg_box.setFont(font)
            
            # Also apply to all buttons
            for button in msg_box.findChildren(QPushButton):
                button.setFont(font)
        
        return msg_box
    
    def show_information(self, title, text):
        """Show information dialog with zoom-scaled font"""
        msg_box = self._create_message_box(QMessageBox.Icon.Information, title, text)
        return msg_box.exec()
    
    def show_warning(self, title, text):
        """Show warning dialog with zoom-scaled font"""
        msg_box = self._create_message_box(QMessageBox.Icon.Warning, title, text)
        return msg_box.exec()
    
    def show_critical(self, title, text):
        """Show critical error dialog with zoom-scaled font"""
        msg_box = self._create_message_box(QMessageBox.Icon.Critical, title, text)
        return msg_box.exec()
    
    def browse_traveler(self):
        """Browse for traveler Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Traveler File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.traveler_input.setText(file_path)
    
    def authenticate_drive(self):
        """Authenticate with Google Drive and Sheets"""
        if not GOOGLE_API_AVAILABLE:
            self.show_critical(
                "Missing Dependencies",
                "Google Drive & Sheets API libraries are not installed."
            )
            return
        
        try:
            self.status_label.setText("🔐 Authenticating with Google...")
            self.show_information(
                "Google Authentication",
                "A browser window will open for you to log in with your Google account.\n\n"
                "Please grant access to:\n"
                "• Google Drive (read-only)\n"
                "• Google Sheets (read-only)"
            )
            
            # Get credentials path
            base_dir = Path(__file__).parent
            credentials_file = base_dir / "google_credentials.json"
            token_file = base_dir / "google_token.pickle"
            
            if not credentials_file.exists():
                self.show_critical(
                    "Credentials Missing",
                    f"Google OAuth credentials file not found:\n{credentials_file}\n\n"
                    "Please follow the setup instructions in FILE_AUDIT_README.md"
                )
                return
            
            # Authenticate
            creds = None
            if token_file.exists():
                with open(token_file, 'rb') as token:
                    creds = pickle.load(token)
            
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    flow = InstalledAppFlow.from_client_secrets_file(
                        str(credentials_file), SCOPES
                    )
                    creds = flow.run_local_server(port=0)
                
                with open(token_file, 'wb') as token:
                    pickle.dump(creds, token)
            
            # Build both Drive and Sheets services
            self.drive_service = build('drive', 'v3', credentials=creds)
            self.sheets_service = build('sheets', 'v4', credentials=creds)
            
            self.status_label.setText("✅ Authenticated with Google Drive & Sheets")
            self.auth_btn.setEnabled(False)
            self.start_btn.setEnabled(True)
            
            self.show_information(
                "Authentication Successful",
                "You are now authenticated with Google Drive & Sheets.\n\n"
                "You can use Excel files or live Google Sheets URLs.\n\n"
                "⚠️ Important: Ensure these APIs are enabled in Google Cloud Console:\n"
                "• Google Drive API\n"
                "• Google Sheets API\n\n"
                "If you get 403 errors, visit:\n"
                "https://console.cloud.google.com/apis/library"
            )
            
        except Exception as e:
            self.status_label.setText(f"❌ Authentication failed: {str(e)}")
            self.show_critical(
                "Authentication Error",
                f"Failed to authenticate with Google Drive:\n\n{str(e)}"
            )
    
    def start_audit(self):
        """Start the file audit"""
        # Validate inputs
        if not self.traveler_input.text():
            self.show_warning("No Input", "Please enter an Excel file path or Google Sheet URL.")
            return
        
        if not self.drive_service or not self.sheets_service:
            self.show_warning("Not Authenticated", "Please authenticate with Google first.")
            return
        
        # Get selected tabs
        tabs = []
        if self.tab_510.isChecked():
            tabs.append("API-510 Traveler")
        if self.tab_570.isChecked():
            tabs.append("API-570 Traveler")
        if self.tab_tank.isChecked():
            tabs.append("Tank Traveler")
        
        if not tabs:
            self.show_warning("No Tabs Selected", "Please select at least one traveler tab to audit.")
            return
        
        # Clear results
        self.results_table.setRowCount(0)
        self.missing_items.clear()
        
        # Disable controls
        self.start_btn.setEnabled(False)
        self.export_excel_btn.setEnabled(False)
        self.export_csv_btn.setEnabled(False)
        
        # Create and start worker
        parent_folder_url = self.parent_folder_input.text().strip() or None
        self.worker = FileAuditWorker(
            self.drive_service,
            self.sheets_service,
            self.traveler_input.text(),
            tabs,
            self.initials_input.text().strip(),
            parent_folder_url  # PHASE 1: Enable fast folder lookup
        )
        
        self.worker.progress.connect(self.update_status)
        self.worker.entity_progress.connect(self.update_entity_progress)
        self.worker.finished.connect(self.audit_finished)
        self.worker.error.connect(self.audit_error)
        
        self.worker.start()
    
    def update_status(self, message):
        """Update status label"""
        self.status_label.setText(message)
    
    def update_entity_progress(self, current, total, entity_name):
        """Update progress bar"""
        progress = int((current / total) * 100)
        self.progress_bar.setValue(progress)
        self.status_label.setText(f"Auditing {current}/{total}: {entity_name}")
    
    def audit_finished(self, missing_items):
        """Handle audit completion"""
        self.missing_items = missing_items
        self.populate_results_table()
        
        self.progress_bar.setValue(100)
        self.status_label.setText(f"✅ Audit complete. Found {len(missing_items)} missing items.")
        
        self.start_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(len(missing_items) > 0)
        self.export_csv_btn.setEnabled(len(missing_items) > 0)
    
    def audit_error(self, error_msg):
        """Handle audit error"""
        self.status_label.setText("❌ Audit failed")
        self.start_btn.setEnabled(True)
        
        self.show_critical(
            "Audit Error",
            f"An error occurred during the audit:\n\n{error_msg}"
        )
    
    def populate_results_table(self):
        """Populate results table with missing items"""
        self.results_table.setRowCount(len(self.missing_items))
        
        for row, item in enumerate(self.missing_items):
            self.results_table.setItem(row, 0, QTableWidgetItem(str(item['Timestamp'])))
            self.results_table.setItem(row, 1, QTableWidgetItem(str(item['Tab'])))
            self.results_table.setItem(row, 2, QTableWidgetItem(str(item['Row'])))
            
            # Entity column with clickable link if folder URL exists
            entity_item = QTableWidgetItem(item['Entity'])
            folder_url = item.get('FolderUrl', '')
            if folder_url:
                # Store URL as data and style as a link
                entity_item.setData(Qt.ItemDataRole.UserRole, folder_url)
                entity_item.setForeground(QColor("#4a90e2"))  # Blue link color
                entity_item.setToolTip(f"Click to open folder:\n{folder_url}")
                # Make it look clickable
                font = entity_item.font()
                font.setUnderline(True)
                entity_item.setFont(font)
            self.results_table.setItem(row, 3, entity_item)
            
            self.results_table.setItem(row, 4, QTableWidgetItem(item['FolderId']))
            self.results_table.setItem(row, 5, QTableWidgetItem(item['Task']))
            self.results_table.setItem(row, 6, QTableWidgetItem(item['Assignee']))
            self.results_table.setItem(row, 7, QTableWidgetItem(item['Trigger']))
            self.results_table.setItem(row, 8, QTableWidgetItem(item['Result']))
            self.results_table.setItem(row, 9, QTableWidgetItem(item['Notes']))
        
        self.results_table.resizeColumnsToContents()
    
    def on_cell_hovered(self, row, column):
        """Change cursor when hovering over clickable entity cells"""
        if column == 3:  # Entity column
            item = self.results_table.item(row, column)
            if item and item.data(Qt.ItemDataRole.UserRole):
                self.results_table.viewport().setCursor(Qt.CursorShape.PointingHandCursor)
                return
        # Reset cursor for non-clickable cells
        self.results_table.viewport().setCursor(Qt.CursorShape.ArrowCursor)
    
    def on_entity_clicked(self, row, column):
        """Handle clicks on entity names to open Drive folder URLs"""
        # Check if Entity column (column 3) was clicked
        if column == 3:
            item = self.results_table.item(row, column)
            if item:
                folder_url = item.data(Qt.ItemDataRole.UserRole)
                if folder_url:
                    try:
                        webbrowser.open(folder_url)
                    except Exception as e:
                        self.show_warning(
                            "Unable to Open Link",
                            f"Could not open the folder URL:\n{folder_url}\n\nError: {str(e)}"
                        )
    
    def export_to_excel(self):
        """Export results to Excel"""
        if not self.missing_items:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            f"FileAudit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # Use openpyxl directly (simpler and more reliable)
                import openpyxl
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "File Audit Results"
                
                # Add headers
                headers = list(self.missing_items[0].keys())
                for col, header in enumerate(headers, 1):
                    sheet.cell(1, col, header)
                
                # Add data
                for row, item in enumerate(self.missing_items, 2):
                    for col, key in enumerate(headers, 1):
                        sheet.cell(row, col, item[key])
                
                workbook.save(file_path)
                
                self.show_information(
                    "Export Successful",
                    f"Results exported to:\n{file_path}"
                )
            except Exception as e:
                self.show_critical(
                    "Export Error",
                    f"Failed to export to Excel:\n\n{str(e)}"
                )
    
    def export_to_csv(self):
        """Export results to CSV"""
        if not self.missing_items:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            f"FileAudit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if file_path:
            try:
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    if self.missing_items:
                        writer = csv.DictWriter(f, fieldnames=self.missing_items[0].keys())
                        writer.writeheader()
                        writer.writerows(self.missing_items)
                
                self.show_information(
                    "Export Successful",
                    f"Results exported to:\n{file_path}"
                )
            except Exception as e:
                self.show_critical(
                    "Export Error",
                    f"Failed to export to CSV:\n\n{str(e)}"
                )
