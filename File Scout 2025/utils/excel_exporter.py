class ExcelExporter:
    """Enhanced Excel exporter with formatting and multiple export options."""
    def __init__(self, theme='light'):
        self.theme = theme
        if theme == 'dark':
            self.header_bg = '1F1F1F'
            self.header_fg = 'FFFFFF'
            self.alt_row_bg = '2A2A2A'
        else:
            self.header_bg = '4472C4'
            self.header_fg = 'FFFFFF'
            self.alt_row_bg = 'E6E6E6'

    def export_data(self, file_path, headers, data, sheet_name="File Scout Results"):
        """Export data to Excel with formatting."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color=self.header_fg)
                cell.fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for row_idx, row_data in enumerate(data, 2):
                row_fill = PatternFill(start_color=self.alt_row_bg, end_color=self.alt_row_bg, fill_type="solid") if row_idx % 2 == 0 else None
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill

            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            ws.freeze_panes = 'A2'
            wb.save(file_path)
            return True
        except ImportError:
            raise ImportError("The 'openpyxl' library is required for Excel export.\nPlease install it: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Failed to export to Excel: {str(e)}")
