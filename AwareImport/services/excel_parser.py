import os
import openpyxl
from typing import Optional
from models.cml_row import CMLRow
from app.constants import (
    KNOWN_SHEET_NAMES,
    EXPECTED_HEADERS,
    PIPING_COLUMN_MAP,
    HEADER_ROW,
    DATA_START_ROW,
)
from utils.helpers import safe_str, extract_system_name_from_filename, temp_open_workbook


def parse_excel_file(file_path: str) -> tuple[list[CMLRow], str, list[str]]:
    # returns (rows, system_name, errors)
    errors: list[str] = []
    rows: list[CMLRow] = []
    system_name = ""

    if not os.path.exists(file_path):
        return rows, system_name, [f"File not found: {file_path}"]

    try:
        with temp_open_workbook(file_path, data_only=True, read_only=True) as wb:
            # find the best sheet
            sheet = _find_best_sheet(wb, errors)
            if sheet is None:
                return rows, system_name, errors or ["No suitable sheet found"]

            # extract system name from header area
            system_name = _extract_system_name(sheet, file_path)

            # extract technician name from header area
            technician = _extract_technician(sheet)

            # extract material type from D4 selector
            material_type = _extract_material_type(sheet)
            if material_type.startswith("Unrecognized:"):
                raw_label = material_type[len("Unrecognized:"):]
                errors.append(
                    f"WARNING: Unrecognized material type in D4 ('{raw_label}'). "
                    f"Defaulting to Carbon Steel — please verify."
                )
                material_type = "Carbon"

            # detect header row and column mapping
            col_map = _detect_columns(sheet)
            if col_map is None:
                # fallback to default mapping
                col_map = PIPING_COLUMN_MAP

            # extract data rows — read ALL rows into a list first so we can
            # do lookahead without re-parsing.  Column A (index 0) is skipped
            # because it has 1000 pre-populated CML numbers and is not reliable
            # as a stop condition.  Column B (index 1) = CML Location and
            # Column T (index 19) = UT Reading are used to determine valid rows.
            file_mtime = os.path.getmtime(file_path) if os.path.exists(file_path) else 0.0

            all_sheet_rows = list(sheet.iter_rows(min_row=DATA_START_ROW, values_only=True))

            # Determine the last "valid" row index in all_sheet_rows:
            # a valid row has a non-blank Column B (CML Location) OR a non-blank
            # Column T (UT Reading).  We scan forward and use a 5-row lookahead
            # to skip over isolated blank rows.
            _LOOKAHEAD = 5  # rows to peek ahead when a blank B+T row is found

            last_valid_idx = -1
            i = 0
            while i < len(all_sheet_rows):
                row = all_sheet_rows[i]
                location_val = safe_str(row[1] if len(row) > 1 else None)
                reading_val = safe_str(row[19] if len(row) > 19 else None)

                if location_val or reading_val:
                    last_valid_idx = i
                    i += 1
                    continue

                # Both B and T are blank — look ahead up to _LOOKAHEAD rows
                found_valid_ahead = False
                for ahead in range(1, _LOOKAHEAD + 1):
                    if i + ahead >= len(all_sheet_rows):
                        break
                    ahead_row = all_sheet_rows[i + ahead]
                    ahead_loc = safe_str(ahead_row[1] if len(ahead_row) > 1 else None)
                    ahead_rd  = safe_str(ahead_row[19] if len(ahead_row) > 19 else None)
                    if ahead_loc or ahead_rd:
                        found_valid_ahead = True
                        break

                if not found_valid_ahead:
                    # Nothing valid in the next _LOOKAHEAD rows — we're done
                    break

                # There IS something valid ahead; skip this blank row and continue
                i += 1

            # Process all rows up to and including last_valid_idx
            for rel_idx, row in enumerate(all_sheet_rows[:last_valid_idx + 1]):
                row_idx = DATA_START_ROW + rel_idx

                location_val = safe_str(row[1] if len(row) > 1 else None)
                reading_val  = safe_str(row[19] if len(row) > 19 else None)

                # Build row data dict from column map
                row_data = {}
                for col_idx, field_name in col_map.items():
                    cell_val = row[col_idx] if col_idx < len(row) else None
                    row_data[field_name] = safe_str(cell_val)

                # Skip rows that have neither location nor reading
                if not row_data.get("cml_location") and not row_data.get("ut_reading"):
                    continue

                cml_row = CMLRow(
                    source_file=file_path,
                    source_sheet=sheet.title,
                    source_row=row_idx,
                    file_modified=file_mtime,
                    system_name=system_name,
                    inspected_by=technician,
                    material_type=material_type,
                    **row_data,
                )
                rows.append(cml_row)
    except Exception as e:
        errors.append(f"Cannot open workbook: {e}")

    return rows, system_name, errors


def _find_best_sheet(wb: openpyxl.Workbook, errors: list[str]) -> Optional[object]:
    # try known sheet names first
    for name in KNOWN_SHEET_NAMES:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() == name.lower():
                return wb[sheet_name]

    # score all sheets by header match
    best_sheet = None
    best_score = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        score = _score_sheet(sheet)
        if score > best_score:
            best_score = score
            best_sheet = sheet

    if best_score >= 3:  # need at least 3 header matches
        return best_sheet

    # fallback: use first sheet if workbook has sheets
    if wb.sheetnames:
        errors.append(f"No known sheet found; using first sheet '{wb.sheetnames[0]}'")
        return wb[wb.sheetnames[0]]

    return None


def _score_sheet(sheet) -> int:
    # score a sheet by how many expected headers it contains in the header row area
    score = 0
    try:
        for row in sheet.iter_rows(min_row=max(1, HEADER_ROW - 1), max_row=HEADER_ROW + 1, values_only=True):
            for cell in row:
                cell_str = safe_str(cell).lower().strip().rstrip(".")
                for header in EXPECTED_HEADERS:
                    if header.lower().rstrip(".") == cell_str:
                        score += 1
                        break
    except Exception:
        pass
    return score


def _detect_columns(sheet) -> Optional[dict[int, str]]:
    # try to auto-detect column mapping from the header row
    try:
        header_row = None
        for row in sheet.iter_rows(min_row=max(1, HEADER_ROW - 1), max_row=HEADER_ROW + 1, values_only=True):
            cells = [safe_str(c).lower().strip().rstrip(".") for c in row]
            if "cml" in cells:
                header_row = cells
                break

        if header_row is None:
            return None

        # map detected headers to field names
        header_to_field = {
            "cml": "cml",
            "cml location": "cml_location",
            "component type": "component_type",
            "component": "component",
            "od": "od",
            "nom": "nom",
            "c.a": "ca",
            "ca": "ca",
            "t-min": "tmin",
            "tmin": "tmin",
            "mat. spec": "mat_spec",
            "mat spec": "mat_spec",
            "material spec": "mat_spec",
            "grade": "mat_grade",
            "mat. grade": "mat_grade",
            "mat grade": "mat_grade",
            "material grade": "mat_grade",
            "pressure": "pressure",
            "temp": "temp",
            "temperature": "temp",
            "j.e": "je",
            "je": "je",
            "joint efficiency": "je",
            "access": "access",
            "insulation": "insulation",
            "install date": "install_date",
            "installed date": "install_date",
            "status": "status",
            "nde": "nde",
            "ut reading": "ut_reading",
            "inspection notes": "inspection_notes",
            "notes": "inspection_notes",
            "comments": "inspection_notes",
            "inspected by": "inspected_by",
            "technician": "inspected_by",
            "inspector": "inspected_by",
        }

        col_map = {}
        seen_fields = set()
        for idx, cell in enumerate(header_row):
            cell_clean = cell.rstrip(".")
            if cell_clean in header_to_field:
                field = header_to_field[cell_clean]
                # only take the first occurrence of each field (e.g. OD in col E, not col W)
                if field not in seen_fields:
                    col_map[idx] = field
                    seen_fields.add(field)

        if len(col_map) >= 3:
            return col_map
    except Exception:
        pass

    return None


def _extract_technician(sheet) -> str:
    """Extract technician / inspected-by name from the header area (e.g. Piping!J3)."""
    try:
        for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
            for i, cell in enumerate(row):
                cell_str = safe_str(cell).lower()
                if any(kw in cell_str for kw in ("technician", "inspected by", "inspector")):
                    # value is typically in the next column(s)
                    for j in range(i + 1, min(i + 4, len(row))):
                        val = safe_str(row[j])
                        if val:
                            return val
    except Exception:
        pass
    return ""


def _extract_material_type(sheet) -> str:
    """Extract material type from D4 selector (merged D4:F4).

    Returns one of:
      'Carbon'         - Carbon Steel circuit
      'Stainless'      - Stainless Steel circuit
      'Mixed'          - Circuit contains both CS and SS components (e.g. "CS & SS")
      'Unrecognized:<raw>' - D4 has a value but it doesn't match any known type;
                            the caller should surface this to the user.
    """
    try:
        for row in sheet.iter_rows(min_row=4, max_row=4, min_col=4, max_col=6, values_only=True):
            for cell in row:
                raw = safe_str(cell).strip()
                val = raw.lower()
                if not val:
                    continue
                # Check for mixed indicator first (contains signals for both materials)
                has_cs = "carbon" in val or "cs" in val
                has_ss = "stainless" in val or "ss" in val
                if has_cs and has_ss:
                    return "Mixed"
                if has_ss:
                    return "Stainless"
                if has_cs:
                    return "Carbon"
                # Non-empty but unrecognized — flag it for the caller
                return f"Unrecognized:{raw}"
    except Exception:
        pass
    return "Carbon"


def extract_inspection_date(file_path: str) -> str:
    """Extract the UT inspection date from cell K1 (merged K1:M1) of the UT sheet.
    Returns a formatted date string or empty string if not found."""
    from datetime import datetime
    try:
        with temp_open_workbook(file_path, data_only=True, read_only=True) as wb:
            sheet = _find_best_sheet(wb, [])
            if sheet is None:
                return ""
            # K1 = column 11, row 1 — read row 1, columns K-M
            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=11, max_col=13, values_only=True):
                for cell in row:
                    val = safe_str(cell).strip()
                    if val:
                        # Try to format if it's a datetime object or parseable date
                        if hasattr(cell, 'strftime'):
                            return cell.strftime("%m/%d/%Y")
                        # Try common date formats
                        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y",
                                    "%d/%m/%Y", "%d/%m/%y"):
                            try:
                                dt = datetime.strptime(val, fmt)
                                return dt.strftime("%m/%d/%Y")
                            except ValueError:
                                continue
                        return val  # return raw string if can't parse
    except Exception:
        pass
    return ""


def _normalize_system_name(name: str) -> str:
    """Normalize system name: replace spaces between letter and number groups with dashes.
    E.g. 'AMINE 008' -> 'AMINE-008', 'COND 005' -> 'COND-005'.
    Already-dashed names like 'AMINE-007' pass through unchanged."""
    import re  # local import to keep module-level imports unchanged
    # Replace space(s) between a letter group and a digit group with a dash
    return re.sub(r'([A-Za-z])\s+(\d)', r'\1-\2', name.strip())


def _extract_system_name(sheet, file_path: str) -> str:
    # try to find system name from header area (rows 1-4)
    try:
        for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
            for i, cell in enumerate(row):
                cell_str = safe_str(cell).lower()
                # look for "pipe circuit name" label, value is usually next cell(s)
                if "circuit name" in cell_str or "pipe circuit" in cell_str:
                    # value is typically in the next column(s)
                    for j in range(i + 1, min(i + 4, len(row))):
                        val = safe_str(row[j])
                        if val:
                            return _normalize_system_name(val)
    except Exception:
        pass

    # fallback to filename
    filename = os.path.basename(file_path)
    return extract_system_name_from_filename(filename)
