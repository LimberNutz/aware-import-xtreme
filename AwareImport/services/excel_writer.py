import os
import openpyxl
from collections import defaultdict
from models.cml_row import CMLRow, EntityInfoRow
from app.constants import PIPING_COLUMN_MAP
from services.excel_parser import _find_best_sheet
from utils.helpers import safe_str


# Reverse map: field_name -> Excel column index (0-based)
_FIELD_TO_COL = {v: k for k, v in PIPING_COLUMN_MAP.items()}

# Preview table column index -> CMLRow field name (must match PreviewTableModel.FIELD_MAP)
_PREVIEW_COL_TO_FIELD = {
    0: "cml", 1: "cml_location", 2: "component_type", 3: "component",
    4: "od", 5: "nom", 6: "ca", 7: "tmin", 8: "mat_spec", 9: "mat_grade",
    10: "pressure", 11: "temp", 12: "je", 13: "access", 14: "insulation",
    15: "install_date", 16: "status", 17: "nde", 18: "inspected_by",
}


def write_back_changes(
    rows: list[CMLRow],
    changed_cells: set[tuple[int, int]],
    progress_callback=None,
) -> list[str]:
    """Write changed preview-table cells back to source Excel workbooks.

    Args:
        rows: The full list of CMLRow objects from the preview model.
        changed_cells: Set of (row_index, preview_col_index) that were edited.

    Returns:
        List of error messages (empty on full success).
    """
    errors: list[str] = []

    # Group changes by (source_file, source_sheet)
    # value: list of (excel_row, excel_col_0based, new_value)
    file_changes: dict[tuple[str, str], list[tuple[int, int, str]]] = defaultdict(list)

    for row_idx, col_idx in changed_cells:
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        if not row.source_file or not row.source_sheet or row.source_row == 0:
            errors.append(
                f"Row {row_idx} ({row.system_name}/{row.cml}): missing source info, cannot write back"
            )
            continue

        field_name = _PREVIEW_COL_TO_FIELD.get(col_idx)
        if not field_name:
            continue

        excel_col = _FIELD_TO_COL.get(field_name)
        if excel_col is None:
            errors.append(
                f"Row {row_idx} ({row.cml}): field '{field_name}' has no Excel column mapping"
            )
            continue

        new_value = getattr(row, field_name, "")
        key = (row.source_file, row.source_sheet)
        file_changes[key].append((row.source_row, excel_col, new_value))

    # Apply changes per workbook
    items = list(file_changes.items())
    total = len(items)
    for file_idx, ((file_path, sheet_name), changes) in enumerate(items):
        if progress_callback:
            progress_callback(file_idx + 1, total)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]

            for excel_row, excel_col_0, new_value in changes:
                # openpyxl cell() is 1-indexed for both row and column
                ws.cell(row=excel_row, column=excel_col_0 + 1, value=new_value)

            wb.save(file_path)
            wb.close()
        except Exception as e:
            errors.append(f"Error writing to {file_path} [{sheet_name}]: {e}")

    return errors


# TA column name -> Excel column index (0-based) for Thickness Activity write-back
_TA_COL_TO_EXCEL = {
    "CML": 0,              # A
    "CML Location": 1,     # B
    "Component Type": 2,   # C
    "Component": 3,        # D
    "OD": 4,               # E
    "Nom.": 5,             # F
    "C.A.": 6,             # G
    "T-Min": 7,            # H
    "Mat. Spec.": 8,       # I
    "Mat. Grade": 9,       # J
    "Pressure": 10,        # K
    "Temp.": 11,           # L
    # M-P (12-15) are blank placeholders, not in the Excel source
    "UT Reading": 19,       # T
    "Inspection Notes": 20, # U
}

# TA view column index -> TA column name
_TA_VIEW_COL_TO_NAME = {
    0: "CML", 1: "CML Location", 2: "Component Type", 3: "Component",
    4: "OD", 5: "Nom.", 6: "C.A.", 7: "T-Min", 8: "Mat. Spec.", 9: "Mat. Grade",
    10: "Pressure", 11: "Temp.",
    12: "First Insp Date", 13: "First UT Reading", 14: "Last Insp Date", 15: "Last UT Reading",
    16: "UT Reading", 17: "Inspection Notes", 18: "Inspected By",
}


# Entity view column index -> EntityInfoRow attribute name (must match EntityInfoModel.FIELD_MAP)
_ENTITY_VIEW_COL_TO_FIELD = {
    4: "equipment_description",
    6: "year_built",
    8: "in_service_date",
    9: "class_name",
    10: "stress_table_used",
    11: "pid_drawing",
    12: "pid_number",
    13: "pfd",
    14: "pfd_number",
    16: "diameter",
    17: "process_service",
}

# Entity field name -> workbook header labels used to locate the value cell
_ENTITY_FIELD_TO_LABELS = {
    "equipment_description": ("description", "circuit description"),
    "pid_number": ("p&id page", "p&id page number", "pid number", "pid page number"),
}


def _find_value_cell(sheet, labels: tuple[str, ...], max_row: int = 4):
    """Find the cell that holds the value adjacent to a label in the header rows.

    Returns (row, column) 1-indexed, or None if the label is not found.
    """
    for row_idx in range(1, max_row + 1):
        max_col = sheet.max_column or 20
        for col_idx in range(1, max_col + 1):
            cell_val = safe_str(sheet.cell(row=row_idx, column=col_idx).value)
            lowered = cell_val.lower().strip().rstrip(":")
            if any(label in lowered for label in labels):
                # Return the first non-empty cell to the right (matching read logic),
                # or the immediately adjacent cell if all are empty.
                for next_col in range(col_idx + 1, min(col_idx + 5, max_col + 2)):
                    existing = safe_str(sheet.cell(row=row_idx, column=next_col).value).strip()
                    if existing:
                        return row_idx, next_col
                return row_idx, col_idx + 1
    return None


def write_back_entity_changes(
    rows: list[EntityInfoRow],
    changed_cells: set[tuple[int, int]],
    progress_callback=None,
) -> list[str]:
    """Write changed Info Page Builder cells back to source Excel workbooks.

    Only fields with known workbook header locations (equipment_description,
    pid_number) are written.  Other editable fields are flagged as INFO.

    Returns:
        List of error/info messages (empty on full success).
    """
    errors: list[str] = []
    skipped_fields: set[str] = set()

    # Group changes by source file: {file_path: {field_name: new_value}}
    file_changes: dict[str, dict[str, str]] = defaultdict(dict)

    for row_idx, col_idx in changed_cells:
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        if not row.source_file or not os.path.exists(row.source_file):
            errors.append(
                f"Row {row_idx} ({row.system_name}): no source file, cannot write back"
            )
            continue

        field_name = _ENTITY_VIEW_COL_TO_FIELD.get(col_idx)
        if not field_name:
            continue

        if field_name not in _ENTITY_FIELD_TO_LABELS:
            skipped_fields.add(field_name)
            continue

        new_value = getattr(row, field_name, "")
        file_changes[row.source_file][field_name] = new_value

    # Apply changes per workbook
    items = list(file_changes.items())
    total = len(items)
    for file_idx, (file_path, field_values) in enumerate(items):
        if progress_callback:
            progress_callback(file_idx + 1, total)
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = _find_best_sheet(wb, [])
            if sheet is None:
                errors.append(f"No suitable sheet in {os.path.basename(file_path)}")
                wb.close()
                continue

            for field_name, new_value in field_values.items():
                labels = _ENTITY_FIELD_TO_LABELS[field_name]
                cell_loc = _find_value_cell(sheet, labels)
                if cell_loc:
                    sheet.cell(row=cell_loc[0], column=cell_loc[1], value=new_value)
                else:
                    errors.append(
                        f"{os.path.basename(file_path)}: '{field_name}' label not found in header"
                    )

            wb.save(file_path)
            wb.close()
        except Exception as e:
            errors.append(f"Error writing to {os.path.basename(file_path)}: {e}")

    if skipped_fields:
        pretty = ", ".join(sorted(f.replace('_', ' ').title() for f in skipped_fields))
        errors.append(
            f"INFO: These fields have no Excel header location and were not written back: {pretty}"
        )

    return errors


def write_back_ta_changes(
    rows: list[dict],
    changed_cells: set[tuple[int, int]],
    progress_callback=None,
) -> list[str]:
    """Write changed Thickness Activity cells back to source Excel workbooks.

    Args:
        rows: The full list of TA row dicts from ThicknessActivityModel.
        changed_cells: Set of (row_index, view_col_index) that were edited.

    Returns:
        List of error messages (empty on full success).
    """
    errors: list[str] = []

    file_changes: dict[tuple[str, str], list[tuple[int, int, str]]] = defaultdict(list)

    for row_idx, col_idx in changed_cells:
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        source_file = row.get("_source_file", "")
        source_sheet = row.get("_source_sheet", "")
        source_row = row.get("_source_row", 0)

        if not source_file or not source_sheet or source_row == 0:
            errors.append(
                f"Row {row_idx} ({row.get('CML', '?')}): missing source info, cannot write back"
            )
            continue

        col_name = _TA_VIEW_COL_TO_NAME.get(col_idx)
        if not col_name:
            continue

        excel_col = _TA_COL_TO_EXCEL.get(col_name)
        if excel_col is None:
            # columns like First Insp Date, Inspected By have no Excel source column
            errors.append(
                f"Row {row_idx} ({row.get('CML', '?')}): column '{col_name}' has no Excel source mapping"
            )
            continue

        new_value = row.get(col_name, "")
        key = (source_file, source_sheet)
        file_changes[key].append((source_row, excel_col, new_value))

    items = list(file_changes.items())
    total = len(items)
    for file_idx, ((file_path, sheet_name), changes) in enumerate(items):
        if progress_callback:
            progress_callback(file_idx + 1, total)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]

            for excel_row, excel_col_0, new_value in changes:
                ws.cell(row=excel_row, column=excel_col_0 + 1, value=new_value)

            wb.save(file_path)
            wb.close()
        except Exception as e:
            errors.append(f"Error writing to {file_path} [{sheet_name}]: {e}")

    return errors
