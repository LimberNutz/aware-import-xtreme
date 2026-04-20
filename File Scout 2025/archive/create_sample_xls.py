#!/usr/bin/env python3
"""
Create sample .xls and .xlsm files to test Excel preview functionality.
"""

import os
from pathlib import Path

def create_sample_excel_files():
    """Create sample .xls and .xlsm files for testing."""
    
    test_dir = Path("preview_test_files")
    test_dir.mkdir(exist_ok=True)
    
    print("📊 Creating sample Excel files for testing...")
    
    # Create sample .xls file using xlrd (we'll create it with xlwt for compatibility)
    try:
        import xlwt
        
        # Create .xls file
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sample Data')
        
        # Add headers
        headers = ['Product', 'Category', 'Price', 'Stock', 'Date Added']
        for col, header in enumerate(headers):
            ws.write(0, col, header)
        
        # Add sample data
        data = [
            ['Laptop', 'Electronics', 999.99, 45, '2023-01-15'],
            ['Mouse', 'Electronics', 25.50, 120, '2023-01-16'],
            ['Keyboard', 'Electronics', 75.00, 85, '2023-01-17'],
            ['Monitor', 'Electronics', 299.99, 30, '2023-01-18'],
            ['Desk Chair', 'Furniture', 199.95, 15, '2023-01-19'],
            ['Notebook', 'Office', 12.99, 200, '2023-01-20'],
            ['Pen Set', 'Office', 8.50, 150, '2023-01-21'],
            ['Phone Case', 'Accessories', 15.99, 75, '2023-01-22'],
            ['USB Cable', 'Electronics', 9.99, 90, '2023-01-23'],
            ['Coffee Mug', 'Kitchen', 6.95, 60, '2023-01-24'],
            ['Backpack', 'Accessories', 49.99, 25, '2023-01-25'],
            ['Headphones', 'Electronics', 89.99, 40, '2023-01-26'],
            ['Stapler', 'Office', 14.99, 35, '2023-01-27'],
            ['Water Bottle', 'Sports', 18.99, 55, '2023-01-28'],
            ['Calculator', 'Office', 22.50, 20, '2023-01-29']
        ]
        
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data):
                ws.write(row_idx, col_idx, value)
        
        xls_file = test_dir / "sample_data.xls"
        wb.save(xls_file)
        print(f"✅ Created .xls file: {xls_file}")
        
    except ImportError:
        print("❌ xlwt not available - cannot create .xls file")
        print("   Install with: pip install xlwt")
    except Exception as e:
        print(f"❌ Error creating .xls file: {e}")
    
    # Create sample .xlsm file (Excel with macros) using openpyxl
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Add headers with formatting
        headers = ['Item ID', 'Description', 'Category', 'Unit Price', 'Quantity', 'Total Value']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add sample data
        data = [
            ['INV001', 'Wireless Mouse', 'Electronics', 25.99, 50, '=D2*E2'],
            ['INV002', 'USB Hub', 'Electronics', 15.50, 75, '=D3*E3'],
            ['INV003', 'Laptop Stand', 'Accessories', 35.00, 30, '=D4*E4'],
            ['INV004', 'Webcam HD', 'Electronics', 79.99, 25, '=D5*E5'],
            ['INV005', 'Desk Lamp', 'Furniture', 45.95, 20, '=D6*E6'],
            ['INV006', 'External HDD', 'Electronics', 89.99, 15, '=D7*E7'],
            ['INV007', 'Bluetooth Speaker', 'Electronics', 55.00, 35, '=D8*E8'],
            ['INV008', 'Phone Charger', 'Electronics', 19.99, 60, '=D9*E9'],
            ['INV009', 'Cable Organizer', 'Accessories', 12.50, 80, '=D10*E10'],
            ['INV010', 'Monitor Arm', 'Furniture', 125.00, 10, '=D11*E11']
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add summary section
        ws.cell(row=14, column=4, value="TOTAL:")
        ws.cell(row=14, column=6, value="=SUM(F2:F11)")
        
        xlsm_file = test_dir / "sample_macro.xlsm"
        wb.save(xlsm_file)
        wb.close()
        print(f"✅ Created .xlsm file: {xlsm_file}")
        
    except ImportError:
        print("❌ openpyxl not available - cannot create .xlsm file")
    except Exception as e:
        print(f"❌ Error creating .xlsm file: {e}")
    
    print(f"\n🎉 Sample Excel files created in: {test_dir.absolute()}")
    print("\nNow you can test:")
    print("• .xls files (old Excel format) - uses xlrd library")
    print("• .xlsx files (modern format) - uses openpyxl library") 
    print("• .xlsm files (Excel with macros) - uses openpyxl library")
    
    return test_dir

if __name__ == "__main__":
    create_sample_excel_files()
